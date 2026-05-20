import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.patches import Ellipse


plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Arial Unicode MS"]
plt.rcParams["axes.unicode_minus"] = False


def add_bg(ax):
    blobs = [
        (0.45, 0.56, 0.56, 0.28),
        (0.59, 0.50, 0.36, 0.21),
        (0.34, 0.48, 0.25, 0.16),
        (0.72, 0.44, 0.22, 0.13),
    ]
    for x, y, w, h in blobs:
        ax.add_patch(
            Ellipse(
                (x, y),
                w,
                h,
                transform=ax.transAxes,
                facecolor="#c9ced3",
                edgecolor="none",
                alpha=0.22,
                zorder=0,
            )
        )


def load_monthly_means(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for d, spread_0509, spread_0105 in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        if d is None:
            continue
        if isinstance(d, (int, float)):
            d = openpyxl.utils.datetime.from_excel(d)
        rows.append(
            {
                "date": pd.to_datetime(d),
                "0509": None if spread_0509 is None else float(spread_0509),
                "0105": None if spread_0105 is None else float(spread_0105),
            }
        )
    df = pd.DataFrame(rows)
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month
    df = df[(df["year"] >= 2021) & (df["year"] <= 2025)]
    g = df.groupby(["year", "month"], as_index=False).mean(numeric_only=True)
    return g


def draw_panel(ax, df, value_col, title, colors):
    months = list(range(1, 13))
    for year in sorted(df["year"].unique()):
        sub = df[df["year"] == year].set_index("month").reindex(months)
        ax.plot(
            months,
            sub[value_col].values,
            lw=3.0,
            marker="o",
            ms=5.0,
            color=colors.get(int(year), "#333333"),
            label=str(int(year)),
            zorder=3,
        )
    ax.set_xlim(1, 12)
    ax.set_xticks(months)
    ax.set_xticklabels([f"{m}月" for m in months], fontsize=12)
    ax.set_ylabel("元/吨", fontsize=13)
    ax.set_title(title, fontsize=20, pad=12)
    ax.grid(axis="y", color="#d0d0d0", lw=1.0, alpha=0.8)
    ax.set_axisbelow(True)
    for sp in ["top", "right", "left", "bottom"]:
        ax.spines[sp].set_visible(False)
    ax.tick_params(axis="both", length=0)


def main():
    src = r"D:\中柏MEG\EG数据库\源数据\月差.xlsx"
    df = load_monthly_means(src)

    colors = {
        2021: "#4f81bd",
        2022: "#c0504d",
        2023: "#9bbb59",
        2024: "#8064a2",
        2025: "#f79646",
    }

    fig = plt.figure(figsize=(16, 10), dpi=170)
    fig.patch.set_facecolor("#efefef")

    bg = fig.add_axes([0, 0, 1, 1])
    bg.set_axis_off()
    add_bg(bg)

    fig.text(0.03, 0.945, "4.X EG：月差月均走势（2021-2025）", fontsize=28, fontweight="bold", color="#4a4a4a")
    fig.lines.append(Line2D([0.03, 0.98], [0.905, 0.905], transform=fig.transFigure, color="#4a4a4a", lw=2.4))
    fig.lines.append(Line2D([0.03, 0.12], [0.905, 0.905], transform=fig.transFigure, color="#c81d24", lw=4.2))
    fig.text(0.92, 0.95, "CIEC", fontsize=28, fontweight="bold", color="#5a5a5a")

    ax1 = fig.add_axes([0.07, 0.56, 0.86, 0.27], facecolor="none")
    draw_panel(ax1, df, "0105", "EG 01-05 月差（月均）", colors)

    ax2 = fig.add_axes([0.07, 0.20, 0.86, 0.27], facecolor="none")
    draw_panel(ax2, df, "0509", "EG 05-09 月差（月均）", colors)

    handles = [Line2D([0], [0], color=colors[y], lw=3.0, marker="o", ms=6, label=str(y)) for y in sorted(colors)]
    fig.legend(handles=handles, loc="lower center", bbox_to_anchor=(0.5, 0.08), ncol=5, frameon=False, fontsize=11)

    fig.text(0.50, 0.035, "数据来源：EG数据库《月差.xlsx》；口径：日度数据按自然月取均值", ha="center", fontsize=12.5, color="#222")
    fig.text(0.97, 0.03, "67", ha="right", fontsize=26, color="#222")

    fig.savefig(r"D:\中柏MEG\results\pictures\eg_month_spread_2021_2025.png", dpi=170, bbox_inches="tight")
    fig.savefig(r"D:\中柏MEG\results\pictures\eg_month_spread_2021_2025.svg", bbox_inches="tight")
    print("saved")


if __name__ == "__main__":
    main()
